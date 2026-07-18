from __future__ import annotations

import base64
import hashlib
import json
import re
import ssl
import subprocess
import tempfile
import urllib.parse
import urllib.request
from urllib.error import HTTPError, URLError
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from processar_escalas_2026 import extract_publication

BASE = Path(__file__).resolve().parents[1]
SOURCES = (
    "https://www.marinha.mil.br/cppr/praticagem",
    "https://www.marinha.mil.br/cppr/praticagem_arquivo",
)
MONTHS = {name: index for index, name in enumerate(
    ("janeiro", "fevereiro", "marco", "abril", "maio", "junho", "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"), 1
)}
HEADERS = {
    "ESCALAS": ["ID", "ANO", "MES", "COMPETENCIA", "VERSAO", "DATA", "DIA_SEMANA", "ORDEM", "TRIGRAMA", "NOME_PRATICO", "SITUACAO", "OBSERVACAO", "ARQUIVO_ORIGEM", "CAMINHO_ARQUIVO", "PAGINA_ORIGEM", "DATA_IMPORTACAO"],
    "PUBLICACOES": ["ID_PUBLICACAO", "ANO", "MES", "COMPETENCIA", "VERSAO", "DATA_PUBLICACAO", "NOME_ARQUIVO", "CAMINHO_LOCAL", "URL_ORIGEM", "HASH_SHA256", "TAMANHO_BYTES", "DATA_DOWNLOAD", "STATUS", "OBSERVACAO"],
    "ALTERACOES": ["ID_ALTERACAO", "COMPETENCIA", "VERSAO_ANTERIOR", "VERSAO_NOVA", "DATA", "ORDEM", "TRIGRAMA", "NOME_PRATICO", "TIPO_ALTERACAO", "VALOR_ANTERIOR", "VALOR_NOVO", "ARQUIVO_ANTERIOR", "ARQUIVO_NOVO", "DATA_COMPARACAO"],
    "PENDENCIAS": ["ID_PENDENCIA", "COMPETENCIA", "VERSAO", "ARQUIVO", "PAGINA", "TRECHO_ORIGINAL", "CAMPO_AFETADO", "MOTIVO", "SUGESTAO_DE_REVISAO", "STATUS_REVISAO"],
}
KNOWN_URLS = (
    "https://assets.marinha.mil.br/cppr/sites/www.marinha.mil.br.cppr/files/2026-03/EscalaDeRodizio_ZP17_1-2026.pdf",
    "https://assets.marinha.mil.br/cppr/sites/www.marinha.mil.br.cppr/files/EscalaDeRodizio_ZP17_3-2026.pdf",
    "https://assets.marinha.mil.br/cppr/sites/www.marinha.mil.br.cppr/files/EscalaDeRodizio_ZP17_3-2026%20ALT%202.pdf",
    "https://assets.marinha.mil.br/cppr/sites/www.marinha.mil.br.cppr/files/EscalaDeRodizio_ZP17_4-2026.pdf",
    "https://assets.marinha.mil.br/cppr/sites/www.marinha.mil.br.cppr/files/EscalaDeRodizio_ZP17_4-2026%20ALT_01.pdf",
    "https://assets.marinha.mil.br/cppr/sites/www.marinha.mil.br.cppr/files/EscalaDeRodizio_ZP17_5-2026.pdf",
    "https://assets.marinha.mil.br/cppr/sites/www.marinha.mil.br.cppr/files/EscalaDeRodizio_ZP17_5-2026%20-%20ALT_1.pdf",
    "https://assets.marinha.mil.br/cppr/sites/www.marinha.mil.br.cppr/files/EscalaDeRodizio_ZP17_5-2026%20-%20ALT_2.pdf",
    "https://assets.marinha.mil.br/cppr/sites/www.marinha.mil.br.cppr/files/EscalaDeRodizio_ZP17_6-2026.pdf",
    "https://assets.marinha.mil.br/cppr/sites/www.marinha.mil.br.cppr/files/EscalaDeRodizio_ZP17_6-2026%20-%20ALT_1.pdf",
    "https://assets.marinha.mil.br/cppr/sites/www.marinha.mil.br.cppr/files/EscalaDeRodizio_ZP17_7-2026.pdf",
    "https://assets.marinha.mil.br/cppr/sites/www.marinha.mil.br.cppr/files/EscalaDeRodizio_ZP17_7-2026%20-%20ALT_1.pdf",
)


class Links(HTMLParser):
    def __init__(self):
        super().__init__()
        self.links, self.href, self.text = [], None, []

    def handle_starttag(self, tag, attrs):
        if tag == "a":
            self.href = dict(attrs).get("href")
            self.text = []

    def handle_data(self, data):
        if self.href:
            self.text.append(data)

    def handle_endtag(self, tag):
        if tag == "a" and self.href:
            self.links.append((self.href, " ".join(self.text)))
            self.href, self.text = None, []


def fetch(url: str) -> bytes:
    host = urllib.parse.urlparse(url).hostname
    if host not in {"www.marinha.mil.br", "assets.marinha.mil.br"}:
        raise ValueError(f"Dominio oficial nao permitido: {host}")
    request = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126 Safari/537.36",
        "Accept": "text/html,application/pdf,application/xhtml+xml,*/*;q=0.8",
        "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.7",
    })
    try:
        with urllib.request.urlopen(request, timeout=90, context=ssl.create_default_context()) as response:
            return response.read()
    except URLError as error:
        if "CERTIFICATE_VERIFY_FAILED" not in str(error):
            raise
        # O certificado continua validado pelo cliente nativo do sistema, que
        # reconhece a cadeia usada pelo host oficial de arquivos da Marinha.
        result = subprocess.run(
            ["curl", "--fail", "--silent", "--show-error", "--location", url],
            check=True, capture_output=True, timeout=90,
        )
        return result.stdout


def clean(value: str) -> str:
    return value.lower().replace("Ã§", "c").replace("Ã¡", "a").replace("Ã£", "a").replace("Ã³", "o").replace("Ã­", "i")


def identify(label: str, url: str):
    text = clean(urllib.parse.unquote(f"{label} {url}"))
    if "escala" not in text or "rodizio" not in text or not re.search(r"zp\s*-?\s*17", text):
        return None
    year_match = re.search(r"(?<!\d)(20\d{2})(?!\d)", text)
    if not year_match or int(year_match.group(1)) < 2026:
        return None
    year = int(year_match.group(1))
    month = next((number for name, number in MONTHS.items() if name in text), None)
    if month is None:
        numeric = re.search(r"zp\s*-?\s*17[_\s-]+(1[0-2]|[1-9])[-_/ ]+20\d{2}", text)
        month = int(numeric.group(1)) if numeric else None
    if month is None:
        return None
    alt = re.search(r"alt(?:eracao)?[_\s-]*0*(\d+)", text)
    version = f"ALT-{int(alt.group(1)):02d}" if alt else "ORIGINAL"
    return year, month, version


def discover():
    found = {}
    for page in SOURCES:
        parser = Links()
        try:
            parser.feed(fetch(page).decode("utf-8", errors="replace"))
        except (HTTPError, URLError):
            continue
        for href, label in parser.links:
            url = urllib.parse.urljoin(page, href)
            item = identify(label, url)
            if item:
                found[(item[0], item[1], item[2], url)] = {"year": item[0], "month": item[1], "version": item[2], "url": url, "label": " ".join(label.split())}
    for url in KNOWN_URLS:
        item = identify("Escala de Rodizio Unica ZP-17", url)
        if item:
            found[(item[0], item[1], item[2], url)] = {"year": item[0], "month": item[1], "version": item[2], "url": url, "label": "Documento oficial confirmado"}
    probe_official_names(found)
    return sorted(found.values(), key=lambda item: (item["year"], item["month"], version_rank(item["version"]), item["url"]))


def probe_official_names(found):
    now = datetime.now()
    periods = [(now.year, now.month)]
    periods.append((now.year + 1, 1) if now.month == 12 else (now.year, now.month + 1))
    base = "https://assets.marinha.mil.br/cppr/sites/www.marinha.mil.br.cppr/files/"
    for year, month in periods:
        if year < 2026:
            continue
        names = [f"EscalaDeRodizio_ZP17_{month}-{year}.pdf"]
        for alt in range(1, 11):
            names.extend((
                f"EscalaDeRodizio_ZP17_{month}-{year} - ALT_{alt}.pdf",
                f"EscalaDeRodizio_ZP17_{month}-{year} ALT_{alt:02d}.pdf",
                f"EscalaDeRodizio_ZP17_{month}-{year} ALT {alt}.pdf",
            ))
        for name in names:
            url = base + urllib.parse.quote(name)
            key_item = identify("Escala de Rodizio Unica ZP-17", url)
            if not key_item or any(key[:3] == key_item for key in found):
                continue
            try:
                data = fetch(url)
            except (HTTPError, URLError):
                continue
            if data.startswith(b"%PDF"):
                found[(key_item[0], key_item[1], key_item[2], url)] = {"year": key_item[0], "month": key_item[1], "version": key_item[2], "url": url, "label": "PDF oficial localizado por verificacao mensal"}


def version_rank(value):
    return 0 if value == "ORIGINAL" else int(re.search(r"\d+", value).group())


def rows_and_changes(publications, checked_at):
    scale_rows, publication_rows, pending_rows, changes = [], [], [], []
    eid = pid = qid = aid = 1
    by_competence = {}
    for publication in publications:
        competence = publication["competencia"]
        year, month = map(int, competence.split("-"))
        by_competence.setdefault(competence, []).append(publication)
        for record in publication["registros"]:
            scale_rows.append([f"ESC-{eid:06d}", year, month, competence, publication["versao"], record["data"], record["dia_semana"], record["ordem"], record["trigrama"], record["nome"], record["situacao"], record["observacao"], publication["arquivo"], record["caminho_arquivo"], record["pagina_original"], checked_at]); eid += 1
        meta = publication["meta"]
        publication_rows.append([f"PUB-{pid:04d}", year, month, competence, publication["versao"], publication["data_publicacao"], publication["arquivo"], "", publication["url_origem"], meta["hash"], meta["size"], checked_at, "PROCESSADO_COM_PENDENCIAS" if publication["pendencias"] else "PROCESSADO", f"{len(publication['pendencias'])} pendencia(s)" if publication["pendencias"] else ""]); pid += 1
        for issue in publication["pendencias"]:
            pending_rows.append([f"PEN-{qid:05d}", competence, publication["versao"], publication["arquivo"], issue["pagina"], issue["trecho_original"], issue["campo_afetado"], issue["motivo"], issue["sugestao_de_revisao"], "PENDENTE"]); qid += 1
    years = sorted({int(item["competencia"][:4]) for item in publications})
    for year in years:
        for month in range(1, 13):
            competence = f"{year}-{month:02d}"
            if competence not in by_competence:
                publication_rows.append([f"PUB-{pid:04d}", year, month, competence, "", "", "", "", "", "", 0, checked_at, "NAO_LOCALIZADO", "NAO LOCALIZADO NA FONTE OFICIAL"]); pid += 1
    for competence, versions in by_competence.items():
        versions.sort(key=lambda item: version_rank(item["versao"]))
        for previous, current in zip(versions, versions[1:]):
            key = lambda row: (row["situacao"], row["data"], row["ordem"])
            old, new = {key(row): row for row in previous["registros"]}, {key(row): row for row in current["registros"]}
            for position in sorted(set(old) | set(new)):
                before, after = old.get(position), new.get(position)
                if before and after and before["trigrama"] == after["trigrama"]:
                    continue
                row = after or before
                kind = "INCLUSAO" if not before else "EXCLUSAO" if not after else "MUDANCA_TRIGRAMA"
                changes.append([f"ALT-{aid:05d}", competence, previous["versao"], current["versao"], row["data"], row["ordem"], row["trigrama"], row["nome"], kind, before["trigrama"] if before else "", after["trigrama"] if after else "", previous["arquivo"], current["arquivo"], checked_at]); aid += 1
    return {"ESCALAS": scale_rows, "PUBLICACOES": publication_rows, "ALTERACOES": changes, "PENDENCIAS": pending_rows}


def write_workbook(tables, path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, headers in HEADERS.items():
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        for row in tables[name]:
            sheet.append(row)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="123C4A")
            cell.font = Font(color="FFFFFF", bold=True)
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(42, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_parts(path):
    target = path.with_suffix(path.suffix + ".parts2")
    target.mkdir(parents=True, exist_ok=True)
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    chunks = [encoded[index:index + 20000] for index in range(0, len(encoded), 20000)]
    for old in target.glob("part-*.txt"):
        old.unlink()
    (target / "manifest.txt").write_text(f"{len(chunks)} 20000\n", encoding="ascii")
    for index, chunk in enumerate(chunks):
        (target / f"part-{index:03d}.txt").write_text(chunk, encoding="ascii")


def main():
    checked_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    candidates = discover()
    if not candidates:
        raise RuntimeError("Nenhuma publicacao oficial ZP-17 de 2026 em diante foi localizada")
    publications = []
    with tempfile.TemporaryDirectory(prefix="zp17-") as folder:
        folder = Path(folder)
        for candidate in candidates:
            data = fetch(candidate["url"])
            if not data.startswith(b"%PDF"):
                raise RuntimeError(f"Arquivo oficial nao e PDF: {candidate['url']}")
            filename = urllib.parse.unquote(Path(urllib.parse.urlparse(candidate["url"]).path).name)
            path = folder / filename
            path.write_bytes(data)
            publication = extract_publication(path, candidate["month"], candidate["version"], candidate["url"], checked_at, candidate["year"])
            publication["meta"] = {"hash": hashlib.sha256(data).hexdigest(), "size": len(data)}
            publications.append(publication)
    if not any(item["registros"] for item in publications):
        raise RuntimeError("Os PDFs foram encontrados, mas nenhum registro seguro foi extraido")
    tables = rows_and_changes(publications, checked_at)
    workbook_path = BASE / "PLANILHA" / "Escala_ZP17_2026.xlsx"
    write_workbook(tables, workbook_path)
    write_parts(workbook_path)
    status = {"checked_at": checked_at, "status": "ATUALIZADO", "source": SOURCES[0], "publications": len(publications), "records": len(tables["ESCALAS"]), "years": sorted({item["competencia"][:4] for item in publications})}
    output = BASE / "APP" / "data" / "update-status.json"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(status, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(status, ensure_ascii=False))


if __name__ == "__main__":
    main()
