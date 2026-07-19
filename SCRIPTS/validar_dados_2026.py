from __future__ import annotations

from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parents[1]
WORKBOOK = BASE / "PLANILHA" / "Escala_ZP17_2026.xlsx"
REQUIRED_SHEETS = ("ESCALAS", "PUBLICACOES", "ALTERACOES", "PENDENCIAS")


def sheet_rows(workbook, name: str):
    rows = list(workbook[name].iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in rows[0]] if rows else []
    return [dict(zip(headers, row)) for row in rows[1:] if any(value not in (None, "") for value in row)]


def main():
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Planilha ausente: {WORKBOOK}")
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    missing_sheets = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
    if missing_sheets:
        raise RuntimeError(f"Abas obrigatorias ausentes: {', '.join(missing_sheets)}")

    scales = sheet_rows(workbook, "ESCALAS")
    publications = sheet_rows(workbook, "PUBLICACOES")
    keys = [(row.get("COMPETENCIA"), row.get("VERSAO"), row.get("DATA"), row.get("SITUACAO"), row.get("ORDEM"), row.get("TRIGRAMA")) for row in scales]
    duplicates = [key for key, count in Counter(keys).items() if count > 1]
    outside = [row for row in scales if not str(row.get("ANO", "")).startswith("20") or int(str(row.get("ANO", "0")) or 0) < 2026]
    missing_source = [row for row in scales if not row.get("ARQUIVO_ORIGEM") or not row.get("PAGINA_ORIGEM")]
    invalid_worked_day = [
        row for row in scales
        if (row.get("SITUACAO") == "PRATICO_EM_SERVICO" and row.get("DIA_TRABALHADO") != "SIM")
        or (row.get("SITUACAO") == "PRATICO_EM_PRONTIDAO" and row.get("DIA_TRABALHADO") != "NÃO")
    ]
    invalid_dates = [row for row in scales if not str(row.get("DATA", ""))[:4].isdigit() or int(str(row.get("DATA"))[:4]) < 2026]
    missing_calendar_days = []
    missing_weekdays = []
    groups = {}
    for row in scales:
        try:
            current = date.fromisoformat(str(row.get("DATA", ""))[:10])
        except ValueError:
            continue
        groups.setdefault((row.get("COMPETENCIA"), row.get("VERSAO")), set()).add(current)
    for key, dates in groups.items():
        if not dates:
            continue
        first, last = min(dates), max(dates)
        expected = {first + timedelta(days=offset) for offset in range((last - first).days + 1)}
        gaps = sorted(expected - dates)
        if gaps:
            missing_calendar_days.append((key, gaps))
        if (last - first).days >= 6:
            absent = sorted(set(range(7)) - {item.weekday() for item in dates})
            if absent:
                missing_weekdays.append((key, absent))
    available = sorted({str(row.get("COMPETENCIA")) for row in publications if row.get("STATUS") != "NAO_LOCALIZADO"})
    issues = []
    if duplicates:
        issues.append(f"Registros duplicados: {len(duplicates)}")
    if outside:
        issues.append(f"Registros fora do escopo 2026 em diante: {len(outside)}")
    if invalid_dates:
        issues.append(f"Datas invalidas ou fora do escopo: {len(invalid_dates)}")
    if missing_source:
        issues.append(f"Registros sem origem completa: {len(missing_source)}")
    if missing_calendar_days:
        issues.append(f"Publicacoes com dias ausentes entre a primeira e a ultima data: {len(missing_calendar_days)}")
    if missing_weekdays:
        issues.append(f"Publicacoes sem cobertura dos sete dias da semana: {len(missing_weekdays)}")
    if invalid_worked_day:
        issues.append(f"Registros com classificacao incorreta de dia trabalhado: {len(invalid_worked_day)}")
    report = [
        "# Validacao da base ZP-17", "",
        f"Consulta: {datetime.now().astimezone().isoformat(timespec='seconds')}", "",
        f"- Publicacoes: {len(publications)}",
        f"- Registros: {len(scales)}",
        f"- Competencias disponiveis: {', '.join(available) or 'nenhuma'}",
        f"- Duplicidades exatas: {len(duplicates)}",
        f"- Registros fora do escopo: {len(outside)}",
        f"- Datas invalidas: {len(invalid_dates)}",
        f"- Registros sem arquivo/pagina: {len(missing_source)}", "",
        f"- Publicacoes com lacunas de datas: {len(missing_calendar_days)}",
        f"- Publicacoes sem algum dia da semana: {len(missing_weekdays)}", "",
        f"- Classificacoes incorretas de dia trabalhado: {len(invalid_worked_day)}", "",
        "## Resultado", "",
        "APROVADO" if not issues else "REPROVADO", "",
    ]
    if issues:
        report += ["## Achados", ""] + [f"- {item}" for item in issues]
    target = BASE / "RELATORIOS" / "validacao_2026.md"
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text("\n".join(report), encoding="utf-8")
    print("VALIDACAO", " | ".join(issues) if issues else "sem inconsistencias estruturais")
    return 1 if issues else 0


if __name__ == "__main__":
    raise SystemExit(main())
